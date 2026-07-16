"""Lógica de generación de presupuestos para ArchPresupuesto.

`generar_presupuesto_ia()`, el SYSTEM_PROMPT y la tool 'generar_presupuesto' están
portados tal cual de ../ArchPresupuesto/core/claude_client.py.

`normalizar_capitulos()` y `calcular_totales()` también están portados tal cual del
original: el servidor siempre recalcula importes y totales, no se confía en la
aritmética de la IA.
"""

import os
from io import BytesIO

from anthropic import Anthropic
from dotenv import load_dotenv
from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

load_dotenv()

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
MODEL_NAME = "claude-sonnet-4-5"

client = Anthropic(api_key=ANTHROPIC_API_KEY) if ANTHROPIC_API_KEY else None


SYSTEM_PROMPT = """Eres un arquitecto técnico y aparejador español, experto en la redacción de presupuestos de \
ejecución de obra para estudios de arquitectura.

Debes generar un presupuesto de obra desglosado por capítulos, con partidas realistas, unidades de medida \
correctas y precios de mercado actuales en España, en euros, coherentes con la tipología, superficie, número \
de plantas y calidad de acabados indicados.

Referencia orientativa de coste de ejecución material (PEM) en obra nueva de vivienda, en €/m² construido \
(ajusta al alza o a la baja según tipología y calidad concreta indicada):
- Calidad básica: 700-950 €/m²
- Calidad media: 950-1300 €/m²
- Calidad alta: 1300-1800 €/m²
- Calidad lujo: 1800-2800 €/m²
Para rehabilitación, reduce el rango orientativo entre un 20% y un 40% según el alcance. Para local comercial \
ajusta proporcionalmente a la baja respecto a vivienda de la misma calidad.

Usa los siguientes capítulos estándar, EXACTAMENTE con estos códigos y nombres, y en este orden. Omite un \
capítulo solo si realmente no aplica al tipo de obra (por ejemplo, movimiento de tierras en una rehabilitación \
de interiores que no afecta a cimentación ni al terreno):

01 - Trabajos previos y demoliciones
02 - Movimiento de tierras
03 - Cimentación
04 - Estructura
05 - Cubierta
06 - Cerramientos y particiones
07 - Revestimientos y acabados
08 - Carpintería y vidriería
09 - Instalaciones
10 - Urbanización y jardinería
11 - Gestión de residuos
12 - Seguridad y salud
13 - Control de calidad
14 - Honorarios técnicos (opcional)

Para el capítulo 09 (Instalaciones), incluye únicamente las instalaciones indicadas por el usuario, con \
partidas específicas y diferenciadas para cada una.

Para el capítulo 14 (Honorarios técnicos), inclúyelo siempre como una estimación orientativa (proyecto básico \
y de ejecución, dirección de obra, dirección de ejecución de obra y coordinación de seguridad y salud), \
calculada como un porcentaje razonable sobre el PEM (habitualmente entre un 6% y un 10%), repartida en 2-4 \
partidas diferenciadas.

Cada partida debe tener: descripción clara y técnica, unidad de medida (m², m³, ml, ud, kg, pa, etc.), \
cantidad coherente con la superficie y tipología indicadas, precio unitario de mercado en euros, e importe \
(cantidad × precio unitario, calculado correctamente).

Genera entre 3 y 8 partidas por capítulo (salvo capítulos que no apliquen). La suma de todos los capítulos \
de obra (excluyendo el capítulo 14 de honorarios) debe resultar en un Presupuesto de Ejecución Material \
coherente con la referencia de €/m² indicada arriba para la superficie construida total.

Devuelve el presupuesto EXCLUSIVAMENTE a través de la herramienta 'generar_presupuesto', sin texto adicional \
fuera de la llamada a la herramienta."""


PRESUPUESTO_TOOL = {
    "name": "generar_presupuesto",
    "description": "Registra el presupuesto de obra desglosado por capítulos y partidas.",
    "input_schema": {
        "type": "object",
        "properties": {
            "capitulos": {
                "type": "array",
                "description": "Lista de capítulos del presupuesto, en orden, con código de dos dígitos.",
                "items": {
                    "type": "object",
                    "properties": {
                        "codigo": {
                            "type": "string",
                            "description": "Código de dos dígitos del capítulo, ej. '01'.",
                        },
                        "nombre": {"type": "string"},
                        "partidas": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "descripcion": {"type": "string"},
                                    "unidad": {"type": "string"},
                                    "cantidad": {"type": "number"},
                                    "precio_unitario": {"type": "number"},
                                    "importe": {"type": "number"},
                                },
                                "required": [
                                    "descripcion",
                                    "unidad",
                                    "cantidad",
                                    "precio_unitario",
                                    "importe",
                                ],
                            },
                        },
                    },
                    "required": ["codigo", "nombre", "partidas"],
                },
            }
        },
        "required": ["capitulos"],
    },
}


def construir_prompt_usuario(datos) -> str:
    instalaciones_txt = (
        ", ".join(datos.instalaciones)
        if datos.instalaciones
        else "No se especifican instalaciones concretas (aplica un criterio estándar básico)"
    )
    observaciones_txt = datos.observaciones.strip() if datos.observaciones else "Ninguna"

    return f"""Genera el presupuesto de obra para el siguiente proyecto:

DATOS GENERALES
- Proyecto: {datos.nombre_proyecto}
- Dirección: {datos.direccion}
- Promotor: {datos.promotor}
- Arquitecto: {datos.arquitecto}
- Fecha: {datos.fecha}

CARACTERÍSTICAS DE LA OBRA
- Tipología: {datos.tipologia}
- Superficie construida total: {datos.superficie_construida} m²
- Número de plantas: {datos.num_plantas}
- Calidad de acabados: {datos.calidad_acabados}

INSTALACIONES A PRESUPUESTAR (capítulo 09)
- {instalaciones_txt}

OBSERVACIONES ADICIONALES DEL PROMOTOR/TÉCNICO
- {observaciones_txt}

Genera el presupuesto completo desglosado por capítulos usando la herramienta 'generar_presupuesto'."""


def generar_presupuesto_ia(datos) -> list:
    if client is None:
        raise HTTPException(
            status_code=500,
            detail="No se ha configurado ANTHROPIC_API_KEY en el servidor.",
        )
    try:
        response = client.messages.create(
            model=MODEL_NAME,
            max_tokens=8192,
            system=SYSTEM_PROMPT,
            tools=[PRESUPUESTO_TOOL],
            tool_choice={"type": "tool", "name": "generar_presupuesto"},
            messages=[{"role": "user", "content": construir_prompt_usuario(datos)}],
        )
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Error al generar el presupuesto con Claude: {exc}")

    for block in response.content:
        if block.type == "tool_use" and block.name == "generar_presupuesto":
            return block.input.get("capitulos", [])

    raise HTTPException(status_code=502, detail="Claude no devolvió un presupuesto estructurado válido.")


def normalizar_capitulos(capitulos_raw: list) -> list:
    """Recalcula importes y subtotales en servidor para no depender de la aritmética de la IA."""
    capitulos = []
    for cap in capitulos_raw:
        partidas = []
        subtotal = 0.0
        for p in cap.get("partidas", []):
            cantidad = float(p.get("cantidad", 0) or 0)
            precio_unitario = float(p.get("precio_unitario", 0) or 0)
            importe = round(cantidad * precio_unitario, 2)
            subtotal += importe
            partidas.append({
                "descripcion": p.get("descripcion", ""),
                "unidad": p.get("unidad", ""),
                "cantidad": cantidad,
                "precio_unitario": precio_unitario,
                "importe": importe,
            })
        capitulos.append({
            "codigo": cap.get("codigo", ""),
            "nombre": cap.get("nombre", ""),
            "partidas": partidas,
            "subtotal": round(subtotal, 2),
        })
    return capitulos


def calcular_totales(capitulos: list) -> dict:
    capitulos_obra = [c for c in capitulos if c["codigo"] != "14"]
    capitulo_honorarios = next((c for c in capitulos if c["codigo"] == "14"), None)

    pem = round(sum(c["subtotal"] for c in capitulos_obra), 2)
    gastos_generales = round(pem * 0.13, 2)
    beneficio_industrial = round(pem * 0.06, 2)
    pec = round(pem + gastos_generales + beneficio_industrial, 2)
    iva_obra = round(pec * 0.10, 2)
    total_obra = round(pec + iva_obra, 2)

    honorarios_base = round(capitulo_honorarios["subtotal"], 2) if capitulo_honorarios else 0.0
    iva_honorarios = round(honorarios_base * 0.21, 2)
    total_honorarios = round(honorarios_base + iva_honorarios, 2)

    presupuesto_total = round(total_obra + total_honorarios, 2)

    return {
        "pem": pem,
        "gastos_generales": gastos_generales,
        "beneficio_industrial": beneficio_industrial,
        "pec": pec,
        "iva_obra": iva_obra,
        "total_obra": total_obra,
        "honorarios_base": honorarios_base,
        "iva_honorarios": iva_honorarios,
        "total_honorarios": total_honorarios,
        "presupuesto_total": presupuesto_total,
    }


EURO_FORMAT = "#,##0.00 €"

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _estilar_fila(ws, row: int, columnas: int, font=None, fill=None, border=False):
    for col in range(1, columnas + 1):
        cell = ws.cell(row=row, column=col)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if border:
            cell.border = BORDER


def construir_excel(contexto: dict) -> bytes:
    datos = contexto["datos"]
    capitulos = contexto["capitulos"]
    totales = contexto["totales"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"

    ws.append([f"Presupuesto de obra — {datos['nombre_proyecto']}"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=14)
    ws.append([f"Dirección: {datos['direccion']}"])
    ws.append([f"Promotor: {datos['promotor']}    Arquitecto: {datos['arquitecto']}    Fecha: {datos['fecha']}"])
    ws.append([])

    for cap in capitulos:
        ws.append([f"Capítulo {cap['codigo']} — {cap['nombre']}"])
        ws.cell(row=ws.max_row, column=1).font = BOLD

        ws.append(["Descripción", "Ud", "Cantidad", "Precio unitario (€)", "Importe (€)"])
        _estilar_fila(ws, ws.max_row, 5, font=HEADER_FONT, fill=HEADER_FILL, border=True)

        for p in cap["partidas"]:
            ws.append([p["descripcion"], p["unidad"], p["cantidad"], p["precio_unitario"], p["importe"]])
            row = ws.max_row
            _estilar_fila(ws, row, 5, border=True)
            ws.cell(row=row, column=3).number_format = "#,##0.00"
            ws.cell(row=row, column=4).number_format = EURO_FORMAT
            ws.cell(row=row, column=5).number_format = EURO_FORMAT

        ws.append(["", "", "", "Subtotal capítulo", cap["subtotal"]])
        row = ws.max_row
        _estilar_fila(ws, row, 5, font=BOLD)
        ws.cell(row=row, column=5).number_format = EURO_FORMAT
        ws.append([])

    ws.append(["RESUMEN DEL PRESUPUESTO"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

    resumen = [
        ("Presupuesto de Ejecución Material (PEM)", totales["pem"]),
        ("Gastos Generales (13%)", totales["gastos_generales"]),
        ("Beneficio Industrial (6%)", totales["beneficio_industrial"]),
        ("Presupuesto de Ejecución por Contrata (PEC)", totales["pec"]),
        ("IVA obra (10%)", totales["iva_obra"]),
        ("Total obra (IVA incluido)", totales["total_obra"]),
    ]
    if totales["honorarios_base"] > 0:
        resumen += [
            ("Honorarios técnicos (base)", totales["honorarios_base"]),
            ("IVA honorarios (21%)", totales["iva_honorarios"]),
            ("Total honorarios (IVA incluido)", totales["total_honorarios"]),
        ]
    resumen.append(("PRESUPUESTO TOTAL", totales["presupuesto_total"]))

    for label, value in resumen:
        ws.append([label, "", "", "", value])
        row = ws.max_row
        ws.cell(row=row, column=1).font = BOLD
        ws.cell(row=row, column=5).font = BOLD
        ws.cell(row=row, column=5).number_format = EURO_FORMAT

    for i, width in enumerate([46, 8, 12, 18, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
