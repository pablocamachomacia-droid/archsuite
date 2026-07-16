"""Generación de informes de superficies en Excel y PDF.

Portado de ../ArchSurface/core/report_excel.py y core/report_pdf.py.
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

NAVY = "12213B"
ACCENT = "0EA5B0"
LIGHT_GRAY = "F2F4F7"
WHITE = "FFFFFF"
TEXT = "1F2937"

CLASS_LABELS = {
    "construida": "Superficie construida",
    "util": "Superficie útil",
    None: "Sin clasificar",
}

FIELD_LABELS = {
    "proyecto": "Proyecto",
    "cliente": "Cliente",
    "ubicacion": "Ubicación",
    "arquitecto": "Arquitecto / Estudio",
}


def calcular_totales(spaces: list) -> dict:
    totals = {}
    for space in spaces:
        key = space.get("classification")
        totals[key] = totals.get(key, 0.0) + space["area"]
    grand_total = sum(totals.values())
    return {"por_clasificacion": totals, "total_general": round(grand_total, 2)}


def _header_fill():
    return PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")


def _accent_fill():
    return PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")


def _stripe_fill():
    return PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")


def build_excel(spaces, source_filename, project_info, company_name="ArchSuite", company_subtitle="Informe técnico de superficies"):
    """Devuelve los bytes de un XLSX con el informe de superficies."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Superficies"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 20

    row = 1
    ws.cell(row=row, column=1, value=company_name).font = Font(size=16, bold=True, color=NAVY)
    row += 1
    ws.cell(row=row, column=1, value=company_subtitle).font = Font(size=10, italic=True, color="6B7280")
    row += 2

    project_info = project_info or {}
    info_rows = [
        ("Archivo de origen", source_filename),
        ("Fecha de generación", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    for key, label in FIELD_LABELS.items():
        value = project_info.get(key)
        if value:
            info_rows.append((label, value))

    for label, value in info_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, color=TEXT, size=9.5)
        ws.cell(row=row, column=2, value=value).font = Font(color=TEXT, size=10)
        row += 1

    row += 1
    table_start_row = row

    headers = ["Espacio", "Capa", "Clasificación", "Superficie (m2)"]
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    for i, space in enumerate(spaces):
        stripe = i % 2 == 1
        values = [
            space.get("name", space["id"]),
            space["layer"],
            CLASS_LABELS.get(space.get("classification"), "Sin clasificar"),
            round(space["area"], 2),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = Font(color=TEXT, size=10)
            if col_idx == 4:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "0.00"
            if stripe:
                cell.fill = _stripe_fill()
        row += 1
    last_data_row = row - 1

    if spaces:
        ws.auto_filter.ref = f"A{table_start_row}:D{last_data_row}"

    row += 1
    ws.cell(row=row, column=1, value="Resumen de superficies").font = Font(size=13, bold=True, color=NAVY)
    row += 1

    totales = calcular_totales(spaces)
    totals = totales["por_clasificacion"]

    ws.cell(row=row, column=1, value="Clasificación").font = Font(bold=True, color=WHITE)
    ws.cell(row=row, column=1).fill = _header_fill()
    ws.cell(row=row, column=2, value="Superficie total (m2)").font = Font(bold=True, color=WHITE)
    ws.cell(row=row, column=2).fill = _header_fill()
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    for col in (3, 4):
        ws.cell(row=row, column=col).fill = _header_fill()
    row += 1

    for key in ("construida", "util", None):
        if key in totals:
            ws.cell(row=row, column=1, value=CLASS_LABELS[key]).font = Font(color=TEXT)
            value_cell = ws.cell(row=row, column=2, value=round(totals[key], 2))
            value_cell.number_format = "0.00"
            value_cell.alignment = Alignment(horizontal="right")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1

    ws.cell(row=row, column=1, value="TOTAL GENERAL").font = Font(bold=True, color=WHITE)
    ws.cell(row=row, column=1).fill = _accent_fill()
    total_cell = ws.cell(row=row, column=2, value=totales["total_general"])
    total_cell.font = Font(bold=True, color=WHITE)
    total_cell.number_format = "0.00"
    total_cell.alignment = Alignment(horizontal="right")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    for col in (2, 3, 4):
        ws.cell(row=row, column=col).fill = _accent_fill()

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _bc3_sanitize(text):
    """Limpia un texto para insertarlo en un campo BC3 separado por '|'."""
    if text is None:
        return ""
    return str(text).replace("|", "/").replace("\r", " ").replace("\n", " ").strip()


def build_bc3(spaces, project_data=None):
    """Devuelve los bytes de un archivo .bc3 (FIEBDC-3) con las mediciones.

    Estructura mínima: por cada espacio, un concepto (~C), su descripción
    larga (~D) y su medición (~M); y al final dos conceptos de totales
    (TOTAL_CONSTRUIDA / TOTAL_UTIL) con su medición. `project_data` se
    acepta para mantener la misma firma que build_excel/build_pdf, pero el
    formato mínimo no tiene un campo estándar de metadatos de proyecto a
    este nivel.
    """
    lines = ["~V|FIEBDC-3/2020|"]

    for space in spaces:
        code = _bc3_sanitize(space["id"])
        name = _bc3_sanitize(space.get("name") or space["id"])
        area = round(space.get("area", 0.0), 2)
        lines.append(f"~C|{code}||m2|{name.upper()}|")
        lines.append(f"~D|{code}|{name}|m2|")
        lines.append(f"~M|{code}|\\1\\{area:.2f}\\|")

    totales = calcular_totales(spaces)
    por_clasificacion = totales["por_clasificacion"]
    total_construida = round(por_clasificacion.get("construida", 0.0), 2)
    total_util = round(por_clasificacion.get("util", 0.0), 2)

    lines.append("~C|TOTAL_CONSTRUIDA||m2|Total superficie construida|")
    lines.append(f"~M|TOTAL_CONSTRUIDA|\\1\\{total_construida:.2f}\\|")
    lines.append("~C|TOTAL_UTIL||m2|Total superficie útil|")
    lines.append(f"~M|TOTAL_UTIL|\\1\\{total_util:.2f}\\|")

    content = "\r\n".join(lines) + "\r\n"
    return content.encode("latin-1", errors="replace")


def _footer(canvas, doc, company_name):
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor("#6b7280"))
    canvas.setLineWidth(0.5)
    canvas.line(2 * cm, 1.6 * cm, A4[0] - 2 * cm, 1.6 * cm)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#6b7280"))
    canvas.drawString(2 * cm, 1.2 * cm, f"{company_name} - Generado con ArchSuite / ArchSurface")
    canvas.drawRightString(A4[0] - 2 * cm, 1.2 * cm, f"Página {doc.page}")
    canvas.restoreState()


def build_pdf(spaces, source_filename, project_info, company_name="ArchSuite", company_subtitle="Informe técnico de superficies"):
    """Devuelve los bytes de un PDF con el informe de superficies."""
    navy = colors.HexColor("#12213b")
    mid_gray = colors.HexColor("#6b7280")
    text_color = colors.HexColor("#1f2937")
    light_gray = colors.HexColor("#f2f4f7")
    accent = colors.HexColor("#0ea5b0")
    white = colors.white

    buffer = io.BytesIO()
    doc = BaseDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2.2 * cm,
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")
    doc.addPageTemplates([
        PageTemplate(id="main", frames=[frame], onPage=lambda c, d: _footer(c, d, company_name))
    ])

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleArch", parent=styles["Title"], textColor=navy, fontSize=20, spaceAfter=2)
    subtitle_style = ParagraphStyle("SubtitleArch", parent=styles["Normal"], textColor=mid_gray, fontSize=10.5, spaceAfter=14)
    section_style = ParagraphStyle("SectionArch", parent=styles["Heading2"], textColor=navy, fontSize=13, spaceBefore=14, spaceAfter=8)
    label_style = ParagraphStyle("LabelArch", parent=styles["Normal"], textColor=mid_gray, fontSize=9.5)
    value_style = ParagraphStyle("ValueArch", parent=styles["Normal"], textColor=text_color, fontSize=10.5)
    th_style = ParagraphStyle("TableHeadArch", parent=styles["Normal"], textColor=white, fontName="Helvetica-Bold", fontSize=9.5, leading=12)
    td_style = ParagraphStyle("TableCellArch", parent=styles["Normal"], textColor=text_color, fontSize=9.5, leading=12)
    td_right_style = ParagraphStyle("TableCellRightArch", parent=td_style, alignment=TA_RIGHT)

    story = [Paragraph(company_name, title_style), Paragraph(company_subtitle, subtitle_style)]

    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M")
    info_rows = [
        ("Archivo de origen", source_filename),
        ("Fecha de generación", generated_at),
    ]
    for key, label in FIELD_LABELS.items():
        value = (project_info or {}).get(key)
        if value:
            info_rows.append((label, value))

    info_table_data = [[Paragraph(label, label_style), Paragraph(str(value), value_style)] for label, value in info_rows]
    info_table = Table(info_table_data, colWidths=[4.5 * cm, doc.width - 4.5 * cm])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(info_table)

    story.append(Paragraph("Detalle de espacios", section_style))

    header = ["Espacio", "Clasificación", "Superficie (m²)"]
    table_data = [[Paragraph(h, th_style) for h in header]]
    for space in spaces:
        table_data.append([
            Paragraph(space.get("name") or space["id"], td_style),
            Paragraph(CLASS_LABELS.get(space.get("classification"), "Sin clasificar"), td_style),
            Paragraph(f"{space['area']:.2f}", td_right_style),
        ])

    class_col_width = 5.2 * cm
    area_col_width = 3.4 * cm
    name_col_width = doc.width - class_col_width - area_col_width
    detail_table = Table(table_data, colWidths=[name_col_width, class_col_width, area_col_width], repeatRows=1)
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("LINEBELOW", (0, 0), (-1, 0), 0.75, navy),
        ("LINEBELOW", (0, 1), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
    ]
    for row_idx in range(1, len(table_data)):
        if row_idx % 2 == 0:
            style_commands.append(("BACKGROUND", (0, row_idx), (-1, row_idx), light_gray))
    detail_table.setStyle(TableStyle(style_commands))
    story.append(detail_table)

    story.append(Paragraph("Resumen de superficies", section_style))

    totales = calcular_totales(spaces)
    totals = totales["por_clasificacion"]

    summary_data = [["Clasificación", "Superficie total (m²)"]]
    for key in ("construida", "util", None):
        if key in totals:
            summary_data.append([CLASS_LABELS[key], f"{totals[key]:.2f}"])
    summary_data.append(["TOTAL GENERAL", f"{totales['total_general']:.2f}"])

    summary_table = Table(summary_data, colWidths=[doc.width - 5.5 * cm, 5.5 * cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("LINEBELOW", (0, 1), (-1, -2), 0.4, colors.HexColor("#e5e7eb")),
        ("TEXTCOLOR", (0, 1), (-1, -2), text_color),
        ("BACKGROUND", (0, -1), (-1, -1), accent),
        ("TEXTCOLOR", (0, -1), (-1, -1), white),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 11),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.4 * cm))

    doc.build(story)
    return buffer.getvalue()
